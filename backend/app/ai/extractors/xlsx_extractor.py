import openpyxl
from typing import Dict, Any, List

class XLSXExtractor:
    def extract(self, file_path: str) -> Dict[str, Any]:
        """Extracts structured sheet names, headers, and rows from XLSX files."""
        sheet_summaries: List[str] = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows_text = []
                for row in sheet.iter_rows(values_only=True):
                    filtered_row = [str(val).strip() for val in row if val is not None and str(val).strip()]
                    if filtered_row:
                        rows_text.append(" | ".join(filtered_row))

                if rows_text:
                    sheet_summaries.append(f"--- SHEET: {sheet_name} ---\n" + "\n".join(rows_text[:50]))
            wb.close()
        except Exception as e:
            sheet_summaries.append(f"XLSX extraction note: {str(e)}")

        full_text = "\n\n".join(sheet_summaries)
        return {
            "full_text": full_text,
            "page_count": len(sheet_summaries) or 1,
            "has_text": len(full_text.strip()) > 0
        }

xlsx_extractor = XLSXExtractor()
